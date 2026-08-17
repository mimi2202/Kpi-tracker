"""Excel report: a Summary sheet (totals, RAG chart, department table)
and a KPI Results sheet with the full data. Requires openpyxl.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import PieChart, Reference
from openpyxl.utils import get_column_letter

from .data import RAG_LABELS

HEADER_FILL = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
RAG_FILLS = {
    "ON_TRACK": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
    "AT_RISK": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    "OFF_TRACK": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    "NO_DATA": PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
}


def build_excel(context, organisation_name: str) -> bytes:
    wb = Workbook()
    _build_summary_sheet(wb.active, context, organisation_name)
    _build_results_sheet(wb.create_sheet("KPI Results"), context)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _header_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    return cell


def _build_summary_sheet(ws, context, organisation_name):
    ws.title = "Summary"
    period = context["period"]

    ws["A1"] = organisation_name
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = f"{context['period_type'].title()} Report"
    ws["A2"].font = Font(size=12, color="6B7280")
    ws["A3"] = period.label if period else "No period data available"
    ws["A3"].font = Font(size=11, color="6B7280")

    ws["A5"] = "Total KPIs"
    ws["B5"] = context["total_kpis"]
    ws["A6"] = "Average Achievement"
    ws["B6"] = f"{context['avg_achievement']}%" if context["avg_achievement"] is not None else "N/A"

    # RAG breakdown table — also the chart's data source.
    _header_cell(ws, 8, 1, "Status")
    _header_cell(ws, 8, 2, "Count")

    row = 9
    for key, label in RAG_LABELS.items():
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=context["rag_counts"][key])
        row += 1
    last_rag_row = row - 1

    if context["total_kpis"] > 0:
        chart = PieChart()
        chart.title = "KPI Status Breakdown"
        data = Reference(ws, min_col=2, min_row=8, max_row=last_rag_row)
        cats = Reference(ws, min_col=1, min_row=9, max_row=last_rag_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 12
        ws.add_chart(chart, "D8")

    # Department summary table — dept_scores is a list of plain dicts,
    # computed live in data.py, not read from a model instance.
    dept_row = last_rag_row + 3
    for col, label in enumerate(["Department", "KPIs", "Avg Achievement", "Status"], start=1):
        _header_cell(ws, dept_row, col, label)

    r = dept_row + 1
    for ds in context["dept_scores"]:
        ws.cell(row=r, column=1, value=ds["department_name"])
        ws.cell(row=r, column=2, value=ds["kpi_count"])
        ws.cell(row=r, column=3, value=ds["avg_achievement"])
        status_cell = ws.cell(row=r, column=4, value=ds["rag_status"].replace("_", " ").title())
        if ds["rag_status"] in RAG_FILLS:
            status_cell.fill = RAG_FILLS[ds["rag_status"]]
        r += 1

    if not context["dept_scores"]:
        ws.cell(row=dept_row + 1, column=1, value="No department data for this period.")

    for col, width in zip("ABCD", (28, 12, 18, 14)):
        ws.column_dimensions[col].width = width


def _build_results_sheet(ws, context):
    headers = ["KPI Code", "KPI Name", "Department", "Target", "Actual", "Achievement %", "Status", "Trend"]
    for col, h in enumerate(headers, start=1):
        _header_cell(ws, 1, col, h)

    row = 2
    for res in context["results"]:
        ws.cell(row=row, column=1, value=res.kpi.code)
        ws.cell(row=row, column=2, value=res.kpi.name)
        ws.cell(row=row, column=3, value=res.department.name)
        ws.cell(row=row, column=4, value=float(res.target_value))
        ws.cell(row=row, column=5, value=float(res.actual_value) if res.actual_value is not None else None)
        ws.cell(row=row, column=6, value=float(res.achievement_percentage) if res.achievement_percentage is not None else None)
        status_cell = ws.cell(row=row, column=7, value=res.rag_status.replace("_", " ").title())
        if res.rag_status in RAG_FILLS:
            status_cell.fill = RAG_FILLS[res.rag_status]
        ws.cell(row=row, column=8, value=res.trend_status.replace("_", " ").title() if res.trend_status else "")
        row += 1

    if row == 2:
        ws.cell(row=2, column=1, value="No KPI results for this period.")

    for i, w in enumerate([14, 32, 22, 10, 10, 14, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w