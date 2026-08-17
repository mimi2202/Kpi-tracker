"""Turns an uploaded file into a list of row dicts, plus an optional forced
"kind" when the format itself is unambiguous (the tracker workbook, for example,
is never anything but results, so there's no need to guess from headers).

Returns (rows, forced_kind). forced_kind is None for CSV, plain Excel, and PDF,
the caller runs classify_rows() on those. It's "tracker_results" when the
tracker workbook format is detected.
"""
import csv
import io

from . import tracker_parser


def parse_file(uploaded_file):
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        return _parse_csv(uploaded_file), None
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return _parse_excel_or_tracker(uploaded_file)
    if name.endswith(".pdf"):
        return _parse_pdf(uploaded_file), None
    raise ValueError("Unsupported file type. Use .csv, .xlsx, or .pdf.")


def _parse_csv(file_obj):
    text = file_obj.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [_normalize_row(row) for row in reader]


def _parse_excel_or_tracker(file_obj):
    from openpyxl import load_workbook

    wb = load_workbook(file_obj, data_only=True)

    if tracker_parser.is_tracker_workbook(wb):
        return tracker_parser.parse_tracker_workbook(wb), "tracker_results"

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], None

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    result = []
    for raw in rows[1:]:
        if all(v is None or str(v).strip() == "" for v in raw):
            continue
        row = {headers[i]: raw[i] for i in range(len(headers)) if i < len(raw)}
        result.append(_normalize_row(row))
    return result, None


def _parse_pdf(file_obj):
    """Extracts the first well-formed table found across all pages.
    Only works for PDFs with real embedded text tables, not scanned images.
    """
    import pdfplumber

    result = []
    with pdfplumber.open(file_obj) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                if not table or len(table) < 2:
                    continue
                headers = [str(h).strip() if h else "" for h in table[0]]
                for raw in table[1:]:
                    if all(v is None or str(v).strip() == "" for v in raw):
                        continue
                    row = {headers[i]: raw[i] for i in range(len(headers)) if i < len(raw)}
                    result.append(_normalize_row(row))

    if not result:
        raise ValueError(
            "No table found in this PDF. PDF import only works with text-based tables, "
            "not scanned or photographed pages."
        )
    return result


def _normalize_row(row):
    return {
        (k or "").strip().lower().replace(" ", "_"): (v.strip() if isinstance(v, str) else v)
        for k, v in row.items()
    }